from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from django.views.decorators.csrf import csrf_exempt
from .models import Module, Session, Attendance, Justification
from accounts.models import StudentProfile, TeacherProfile, User
from .forms import SessionForm, ModuleForm, JustificationForm, ManualAttendanceForm
import json
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def teacher_required(func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_teacher():
            messages.error(request, "Accès refusé.")
            return redirect('dashboard:home')
        return func(request, *args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper


def student_required(func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_student():
            messages.error(request, "Accès refusé.")
            return redirect('dashboard:home')
        return func(request, *args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper


def admin_required(func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_admin():
            messages.error(request, "Accès refusé.")
            return redirect('dashboard:home')
        return func(request, *args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper


# ===== TEACHER VIEWS =====

@login_required
@teacher_required
def teacher_modules(request):
    try:
        teacher = request.user.teacher_profile
        modules = Module.objects.filter(teacher=teacher, is_active=True)
    except TeacherProfile.DoesNotExist:
        modules = []
        messages.warning(request, "Profil enseignant non trouvé.")

    return render(request, 'teacher/modules.html', {'modules': modules})


@login_required
@teacher_required
def create_module(request):
    form = ModuleForm()
    if request.method == 'POST':
        form = ModuleForm(request.POST)
        if form.is_valid():
            module = form.save(commit=False)
            module.teacher = request.user.teacher_profile
            module.save()
            messages.success(request, f'Module "{module.name}" créé avec succès!')
            return redirect('attendance:teacher_modules')
    return render(request, 'teacher/create_module.html', {'form': form})


@login_required
@teacher_required
def create_session(request, module_id=None):
    module = None
    if module_id:
        module = get_object_or_404(Module, id=module_id, teacher=request.user.teacher_profile)

    form = SessionForm(teacher=request.user.teacher_profile, initial={'module': module})

    if request.method == 'POST':
        form = SessionForm(request.POST, teacher=request.user.teacher_profile)
        if form.is_valid():
            session = form.save()

            # Auto-create attendance records for enrolled students
            module = session.module
            students = StudentProfile.objects.filter(
                year_of_study=module.year_of_study,
                department=module.department
            )
            for student in students:
                Attendance.objects.get_or_create(
                    student=student,
                    session=session,
                    defaults={'status': 'absent'}
                )

            messages.success(request, f'Séance créée pour le {session.date}!')
            return redirect('attendance:session_detail', pk=session.pk)

    return render(request, 'teacher/create_session.html', {'form': form, 'module': module})


@login_required
@teacher_required
def session_detail(request, pk):
    session = get_object_or_404(Session, pk=pk, module__teacher=request.user.teacher_profile)
    attendances = session.attendances.select_related('student__user').all()

    stats = {
        'total': attendances.count(),
        'present': attendances.filter(status='present').count(),
        'absent': attendances.filter(status='absent').count(),
        'late': attendances.filter(status='late').count(),
        'justified': attendances.filter(status='justified').count(),
    }

    return render(request, 'teacher/session_detail.html', {
        'session': session,
        'attendances': attendances,
        'stats': stats,
    })


@login_required
@teacher_required
def generate_qr(request, session_id):
    session = get_object_or_404(Session, pk=session_id, module__teacher=request.user.teacher_profile)

    expiry = int(request.POST.get('expiry_minutes', 30))
    session.generate_qr_code(expiry_minutes=expiry)

    messages.success(request, f'QR Code généré! Valide pendant {expiry} minutes.')
    return redirect('attendance:session_detail', pk=session_id)


@login_required
@teacher_required
def deactivate_qr(request, session_id):
    session = get_object_or_404(Session, pk=session_id, module__teacher=request.user.teacher_profile)
    session.deactivate_qr()
    messages.info(request, 'QR Code désactivé.')
    return redirect('attendance:session_detail', pk=session_id)


@login_required
@teacher_required
def module_absences(request, module_id):
    module = get_object_or_404(Module, pk=module_id, teacher=request.user.teacher_profile)
    students = StudentProfile.objects.filter(
        year_of_study=module.year_of_study,
        department=module.department
    ).select_related('user')

    sessions = Session.objects.filter(module=module)
    total_sessions = sessions.count()

    student_stats = []
    for student in students:
        absences = Attendance.objects.filter(
            student=student, session__module=module, status__in=['absent']
        ).count()
        presents = Attendance.objects.filter(
            student=student, session__module=module, status='present'
        ).count()
        percentage = round((absences / total_sessions * 100), 2) if total_sessions > 0 else 0
        student_stats.append({
            'student': student,
            'absences': absences,
            'presents': presents,
            'percentage': percentage,
            'is_at_risk': percentage >= 30,
            'is_critical': percentage >= 50,
        })

    return render(request, 'teacher/module_absences.html', {
        'module': module,
        'student_stats': student_stats,
        'total_sessions': total_sessions,
    })

# attendance/views.py — ajouter

@login_required
@teacher_required
def teacher_group_list(request):
    """Enseignant : liste présence/absence par séance."""
    teacher = request.user.teacher_profile
    selected_session_id = request.GET.get('session_id', '')
    export = request.GET.get('export', '')

    sessions = Session.objects.filter(
        module__teacher=teacher
    ).select_related('module').order_by('-date', '-start_time')

    result = []
    selected_session = None

    if selected_session_id:
        selected_session = get_object_or_404(
            Session, pk=selected_session_id, module__teacher=teacher
        )
        attendances = Attendance.objects.filter(
            session=selected_session
        ).select_related('student__user')
        for att in attendances:
            result.append({
                'student': att.student,
                'status': att.get_status_display(),
                'status_code': att.status,
            })

    if export == 'excel' and selected_session:
        return export_group_list_excel(
            result,
            selected_session.module.department,
            selected_session.module.group,
            str(selected_session.date)
        )

    return render(request, 'teacher/group_list.html', {
        'sessions': sessions,
        'result': result,
        'selected_session': selected_session,
        'selected_session_id': selected_session_id,
    })


@login_required
@teacher_required
def mark_attendance_manual(request, session_id):
    session = get_object_or_404(Session, pk=session_id, module__teacher=request.user.teacher_profile)

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        status = request.POST.get('status')

        try:
            student = StudentProfile.objects.get(id=student_id)
            attendance, _ = Attendance.objects.get_or_create(
                student=student,
                session=session,
                defaults={'status': status}
            )
            attendance.status = status
            attendance.marked_by = request.user
            attendance.save()

            # ── Notifications automatiques ──
            from notifications.models import Notification

            if status == 'absent':
                total_sessions = Session.objects.filter(module=session.module).count()
                total_absences = Attendance.objects.filter(
                    student=student,
                    session__module=session.module,
                    status='absent'
                ).count()
                taux = round((total_absences / total_sessions * 100), 1) if total_sessions > 0 else 0

                Notification.objects.create(
                    user=student.user,
                    title=f'Absence enregistrée — {session.module.code}',
                    message=f'Vous avez été marqué absent à la séance du {session.date} pour le module {session.module.name}.',
                    notification_type='absence',
                )

                if taux >= 50:
                    Notification.objects.create(
                        user=student.user,
                        title=f'🔴 CRITIQUE — {session.module.code}',
                        message=f'Votre taux d\'absence dans {session.module.name} a atteint {taux}%. Contactez immédiatement votre administration !',
                        notification_type='alert',
                    )
                elif taux >= 30:
                    Notification.objects.create(
                        user=student.user,
                        title=f'⚠️ Alerte absences — {session.module.code}',
                        message=f'Votre taux d\'absence dans {session.module.name} a atteint {taux}%. Vous risquez de dépasser le seuil autorisé.',
                        notification_type='alert',
                    )

            elif status == 'late':
                Notification.objects.create(
                    user=student.user,
                    title=f'Retard enregistré — {session.module.code}',
                    message=f'Vous avez été marqué en retard à la séance du {session.date} pour le module {session.module.name}.',
                    notification_type='absence',
                )

            return JsonResponse({'success': True, 'status': attendance.get_status_display()})

        except StudentProfile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Étudiant non trouvé'})

    return JsonResponse({'success': False})


# ===== STUDENT VIEWS =====

@login_required
@student_required
def scan_qr(request, token):
    try:
        session = Session.objects.get(qr_token=token)
    except Session.DoesNotExist:
        messages.error(request, 'QR Code invalide.')
        return redirect('dashboard:home')

    if not session.is_qr_valid():
        messages.error(request, 'Ce QR Code est expiré ou désactivé.')
        return redirect('dashboard:home')

    student = request.user.student_profile

    # Check if already scanned
    attendance, created = Attendance.objects.get_or_create(
        student=student,
        session=session,
        defaults={
            'status': 'present',
            'scan_time': timezone.now(),
            'scan_ip': request.META.get('REMOTE_ADDR'),
            'marked_by': request.user,
        }
    )

    if not created:
        if attendance.status == 'present':
            messages.info(request, 'Vous avez déjà été marqué présent pour cette séance.')
        else:
            attendance.status = 'present'
            attendance.scan_time = timezone.now()
            attendance.save()
            messages.success(request, '✅ Présence enregistrée avec succès!')
            # Notification de présence confirmée
        from notifications.models import Notification
        Notification.objects.create(
            user=student.user,
            title=f'✅ Présence confirmée — {session.module.code}',
            message=f'Votre présence a été enregistrée pour la séance du {session.date} ({session.get_session_type_display()}).',
            notification_type='info',
        )
    else:
        messages.success(request, '✅ Présence enregistrée avec succès!')

    return render(request, 'student/scan_success.html', {
        'session': session,
        'attendance': attendance,
    })


@login_required
@student_required
def student_absences(request):
    student = request.user.student_profile
    absences = Attendance.objects.filter(
        student=student,
        status__in=['absent', 'late', 'justified']
    ).select_related('session__module').order_by('-session__date')

    # Stats per module
    modules = Module.objects.filter(
        year_of_study=student.year_of_study,
        department=student.department
    )

    module_stats = []
    for module in modules:
        total = Session.objects.filter(module=module).count()
        absent = Attendance.objects.filter(
            student=student, session__module=module, status='absent'
        ).count()
        percentage = round((absent / total * 100), 2) if total > 0 else 0
        module_stats.append({
            'module': module,
            'total_sessions': total,
            'absences': absent,
            'percentage': percentage,
            'alert': percentage >= 30,
        })

    return render(request, 'student/absences.html', {
        'absences': absences,
        'module_stats': module_stats,
    })


@login_required
@student_required
def justify_absence(request, attendance_id):
    attendance = get_object_or_404(Attendance, pk=attendance_id, student=request.user.student_profile)

    if hasattr(attendance, 'justification'):
        messages.warning(request, 'Cette absence a déjà une justification soumise.')
        return redirect('attendance:student_absences')

    form = JustificationForm()

    if request.method == 'POST':
        form = JustificationForm(request.POST, request.FILES)
        if form.is_valid():
            justification = form.save(commit=False)
            justification.attendance = attendance
            justification.student = request.user.student_profile
            justification.save()
            messages.success(request, 'Justification soumise avec succès! En attente d\'approbation.')
            return redirect('attendance:student_absences')

    return render(request, 'student/justify.html', {
        'attendance': attendance,
        'form': form,
    })


@login_required
@student_required
def student_history(request):
    student = request.user.student_profile
    all_attendances = Attendance.objects.filter(
        student=student
    ).select_related('session__module').order_by('-session__date')

    return render(request, 'student/history.html', {
        'attendances': all_attendances,
    })


# ===== ADMIN VIEWS =====

@login_required
@admin_required
def admin_justifications(request):
    justifications = Justification.objects.select_related(
        'student__user', 'attendance__session__module'
    ).order_by('-submitted_at')

    status_filter = request.GET.get('status', 'pending')
    if status_filter != 'all':
        justifications = justifications.filter(status=status_filter)

    return render(request, 'admin_panel/justifications.html', {
        'justifications': justifications,
        'status_filter': status_filter,
    })


@login_required
@admin_required
def review_justification(request, pk):
    justification = get_object_or_404(Justification, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        comment = request.POST.get('admin_comment', '')

        justification.admin_comment = comment
        justification.reviewed_by = request.user
        justification.reviewed_at = timezone.now()

        if action == 'approve':
            justification.status = 'approved'
            justification.attendance.status = 'justified'
            justification.attendance.save()
            messages.success(request, 'Justification approuvée.')
        elif action == 'reject':
            justification.status = 'rejected'
            messages.warning(request, 'Justification rejetée.')

        justification.save()

        # Create notification for student
        from notifications.models import Notification
        Notification.objects.create(
            user=justification.student.user,
            title='Justification ' + ('approuvée' if action == 'approve' else 'rejetée'),
            message=f'Votre justification pour la séance du {justification.attendance.session.date} a été {justification.get_status_display().lower()}.',
            notification_type='justification',
        )

        return redirect('attendance:admin_justifications')

    return render(request, 'admin_panel/review_justification.html', {
        'justification': justification,
    })


@login_required
@admin_required
def admin_all_absences(request):
    students = StudentProfile.objects.select_related('user').all()
    departments = students.values_list('department', flat=True).distinct()

    # Filtre par département
    dept_filter = request.GET.get('dept', '')
    if dept_filter:
        students = students.filter(department=dept_filter)

    student_stats = []
    for student in students:
        total = Attendance.objects.filter(student=student).count()
        absent = Attendance.objects.filter(student=student, status='absent').count()
        percentage = round((absent / total * 100), 2) if total > 0 else 0
        student_stats.append({
            'student': student,
            'total': total,
            'absent': absent,
            'percentage': percentage,
        })

    # Export Excel
    if request.GET.get('export') == 'excel':
        return export_absences_excel(student_stats)

    # Export CSV
    if request.GET.get('export') == 'csv':
        return export_absences_csv(student_stats)

    return render(request, 'admin_panel/all_absences.html', {
        'student_stats': student_stats,
        'departments': departments,
    })


def export_absences_excel(student_stats):
    """Exporte les stats d'absence en fichier Excel stylisé."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Absences'

    # En-têtes
    headers = ['N° Étudiant', 'Nom complet', 'Département', 'Groupe',
               'Total séances', 'Absences', 'Taux (%)', 'Statut']
    header_fill = PatternFill(start_color='6B1F2A', end_color='6B1F2A', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Données
    for row, item in enumerate(student_stats, 2):
        s = item['student']
        status = 'Critique' if item['percentage'] >= 50 else (
                 'À risque' if item['percentage'] >= 30 else 'Normal')
        ws.append([
            s.student_id,
            s.user.get_full_name(),
            s.department,
            s.group,
            item['total'],
            item['absent'],
            item['percentage'],
            status
        ])
        # Colorier les lignes selon le risque
        if item['percentage'] >= 50:
            fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = fill
        elif item['percentage'] >= 30:
            fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = fill

    # Largeurs automatiques
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=absences.xlsx'
    wb.save(response)
    return response


def export_absences_csv(student_stats):
    """Export CSV simple."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename=absences.csv'
    response.write('\ufeff')  # BOM pour Excel
    writer = csv.writer(response)
    writer.writerow(['N° Étudiant', 'Nom', 'Département', 'Groupe',
                     'Total', 'Absences', 'Taux%'])
    for item in student_stats:
        s = item['student']
        writer.writerow([s.student_id, s.user.get_full_name(), s.department,
                         s.group, item['total'], item['absent'], item['percentage']])
    return response

# attendance/views.py — ajouter cette vue

@login_required
@admin_required
def admin_group_list(request):
    """Liste des étudiants par groupe avec statut présence/absence."""
    from datetime import date as dt_date

    selected_dept = request.GET.get('dept', '')
    selected_group = request.GET.get('group', '')
    selected_date = request.GET.get('date', str(dt_date.today()))
    export = request.GET.get('export', '')

    all_students = StudentProfile.objects.select_related('user').all()
    departments = all_students.values_list('department', flat=True).distinct()
    groups = all_students.values_list('group', flat=True).distinct()

    students = all_students
    if selected_dept:
        students = students.filter(department=selected_dept)
        groups = students.values_list('group', flat=True).distinct()
    if selected_group:
        students = students.filter(group=selected_group)

    # Construire la liste avec statut
    result = []
    for student in students:
        sessions_today = Session.objects.filter(date=selected_date)
        if sessions_today.exists():
            attendance = Attendance.objects.filter(
                student=student,
                session__date=selected_date
            ).first()
            status = attendance.get_status_display() if attendance else 'Absent'
            status_code = attendance.status if attendance else 'absent'
        else:
            status = '—'
            status_code = 'none'
        result.append({
            'student': student,
            'status': status,
            'status_code': status_code,
        })

    # Export Excel de la liste
    if export == 'excel':
        return export_group_list_excel(result, selected_dept, selected_group, selected_date)

    return render(request, 'admin_panel/group_list.html', {
        'result': result,
        'departments': departments,
        'groups': groups,
        'selected_dept': selected_dept,
        'selected_group': selected_group,
        'selected_date': selected_date,
    })


def export_group_list_excel(result, dept, group, date_str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Groupe {group or "Tous"}'
    ws['A1'] = f'Liste de présence — {dept} — Groupe {group} — {date_str}'
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    headers = ['N° Étudiant', 'Nom complet', 'Groupe', 'Statut']
    header_fill = PatternFill(start_color='1E2D45', end_color='1E2D45', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
    for item in result:
        s = item['student']
        ws.append([s.student_id, s.user.get_full_name(), s.group, item['status']])
        if item['status_code'] == 'absent':
            fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).fill = fill
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=groupe_{group}_{date_str}.xlsx'
    wb.save(response)
    return response


@login_required
@admin_required
def import_students_excel(request):
    """Importer une liste d'étudiants depuis Excel."""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            success_count = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]:  # Ignorer lignes vides
                    continue
                try:
                    # Colonnes attendues :
                    # A: student_id, B: first_name, C: last_name,
                    # D: email, E: department, F: year, G: group, H: username
                    student_id, first_name, last_name = row[0], row[1], row[2]
                    email = row[3] or ''
                    department = row[4] or 'Non défini'
                    year = int(row[5] or 1)
                    group = str(row[6] or '')
                    username = row[7] or f'stu_{student_id}'

                    if User.objects.filter(username=username).exists():
                        errors.append(f'Ligne {row_idx}: {username} existe déjà')
                        continue

                    user = User.objects.create_user(
                        username=username,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        password=str(student_id),  # Mot de passe initial = student_id
                        role='student'
                    )
                    StudentProfile.objects.create(
                        user=user,
                        student_id=str(student_id),
                        department=department,
                        year_of_study=year,
                        group=group
                    )
                    success_count += 1
                except Exception as e:
                    errors.append(f'Ligne {row_idx}: {str(e)}')

            messages.success(request, f'{success_count} étudiant(s) importé(s) avec succès.')
            if errors:
                for err in errors[:5]:  # Afficher max 5 erreurs
                    messages.warning(request, err)
        except Exception as e:
            messages.error(request, f'Erreur de lecture du fichier: {str(e)}')

        return redirect('attendance:import_students')

    return render(request, 'admin_panel/import_students.html')

@login_required
@admin_required
def download_template(request):
    """Télécharger le modèle Excel vide."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Étudiants'
    headers = ['N° Étudiant', 'Prénom', 'Nom', 'Email',
               'Département', 'Année (1-5)', 'Groupe', 'Username']
    header_fill = PatternFill(start_color='6B1F2A', end_color='6B1F2A', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
    # Exemple de ligne
    ws.append(['ETU20240001', 'Mohamed', 'Alami', 'malami@univ.ma',
               'Génie Informatique', 1, 'G1', 'malami'])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=modele_etudiants.xlsx'
    wb.save(response)
    return response


@login_required
@admin_required
def admin_student_detail(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)
    absences = Attendance.objects.filter(
        student=student
    ).select_related('session__module').order_by('-session__date')

    return render(request, 'admin_panel/student_detail.html', {
        'student': student,
        'absences': absences,
    })